from flask import render_template, request, redirect, url_for, session, flash, make_response, send_file
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
from database import get_db
from functools import wraps
import pyotp
import qrcode
import base64
import io
import json
import math
import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def no_cache(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        response = make_response(f(*args, **kwargs))
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        return response
    return decorated_function


def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        last_active = session.get('last_active')
        if last_active:
            elapsed = datetime.utcnow() - datetime.fromisoformat(last_active)
            if elapsed > timedelta(minutes=15):
                session.clear()
                flash('You were logged out due to inactivity.', 'warning')
                return redirect(url_for('login'))
        session['last_active'] = datetime.utcnow().isoformat()
        return f(*args, **kwargs)
    return decorated_function


def get_user_by_role(db, role, user_id):
    if role == 'adviser':
        return db.execute('SELECT * FROM advisers WHERE adviser_id = ?', (user_id,)).fetchone()
    return db.execute('SELECT * FROM customers WHERE customer_id = ?', (user_id,)).fetchone()


def get_customer_account(db, account_id, customer_id):
    return db.execute(
        'SELECT * FROM accounts WHERE account_id = ? AND customer_id = ?', 
        (account_id, customer_id)
    ).fetchone()


def _get_market_data():
    today = datetime.utcnow()
    rng = random.Random(int(today.strftime('%Y%m%d')))

    recent_labels, recent_values = [], []
    val = 7580.0
    for i in range(30):
        d = today - timedelta(days=29 - i)
        recent_labels.append(d.strftime('%d %b'))
        val += rng.uniform(-35, 40) + math.sin(i * 0.4) * 12
        recent_values.append(round(val, 2))

    historical_labels, historical_values = [], []
    val = 7050.0
    for i in range(12):
        d = today - timedelta(days=(11 - i) * 30)
        historical_labels.append(d.strftime('%b %Y'))
        val += rng.uniform(-90, 110) + math.sin(i * 0.5) * 40
        historical_values.append(round(val, 2))

    news = [
        {'title': 'FTSE 100 Gains on Easing Inflation Data', 'source': 'Reuters', 'url': '#',
         'summary': 'UK equities advanced as CPI figures came in below forecasts, raising hopes of further Bank of England rate cuts.'},
        {'title': 'Bank of England Signals Gradual Rate Path', 'source': 'BBC News', 'url': '#',
         'summary': 'The Monetary Policy Committee held rates steady while noting improving conditions for a measured easing cycle.'},
        {'title': 'Tech Stocks Lead Broad Market Recovery', 'source': 'Bloomberg', 'url': '#',
         'summary': 'Technology equities surged globally as Q1 earnings broadly beat analyst expectations.'},
        {'title': 'Oil Prices Stabilise Near $82 a Barrel', 'source': 'Financial Times', 'url': '#',
         'summary': 'Brent crude settled on tight supply after OPEC+ confirmed output cuts through mid-year.'},
        {'title': 'Sterling Strengthens on Positive Trade Figures', 'source': 'The Guardian', 'url': '#',
         'summary': 'The pound climbed to a three-month high after UK trade balance data exceeded expectations.'},
    ]

    recent_trend = {
        'labels': recent_labels,
        'values': recent_values,
        'description': 'FTSE 100 index — last 30 trading days.',
    }
    historical_trend = {
        'labels': historical_labels,
        'values': historical_values,
        'description': '12-month FTSE 100 index performance.',
    }
    return recent_trend, historical_trend, news


def register_routes(app):
    
    @app.route('/')
    @no_cache
    def index():
        if 'user_id' not in session: return redirect(url_for('login'))
        db = get_db()
        if session.get('role') == 'customer':
            accounts = db.execute(
                'SELECT * FROM accounts WHERE customer_id = ? ORDER BY account_id DESC',
                (session['user_id'],)
            ).fetchall()
            current_adviser = db.execute('''
                SELECT a.adviser_id, a.name
                FROM user_assignments ua
                JOIN advisers a ON ua.adviser_id = a.adviser_id
                WHERE ua.customer_id = ?
            ''', (session['user_id'],)).fetchone()
            pending_request = db.execute(
                'SELECT request_id FROM customer_requests WHERE customer_id = ?',
                (session['user_id'],)
            ).fetchone()
            categories = db.execute(
                'SELECT * FROM categories WHERE customer_id IS NULL OR customer_id = ? ORDER BY name',
                (session['user_id'],)
            ).fetchall()
            budgets = db.execute('''
                SELECT b.budget_id, b.maximum_amount, b.start_date, b.end_date,
                       b.category_id, c.name AS category_name, c.colour AS category_colour,
                       COALESCE((
                           SELECT SUM(ABS(t.amount))
                           FROM transactions t
                           JOIN accounts a ON t.account_id = a.account_id
                           WHERE a.customer_id = b.customer_id
                             AND t.category_id = b.category_id
                             AND t.amount < 0
                             AND date(t.transaction_date) >= date(b.start_date)
                             AND date(t.transaction_date) <= date(b.end_date)
                       ), 0) AS spent
                FROM budgets b
                LEFT JOIN categories c ON b.category_id = c.category_id
                WHERE b.customer_id = ?
                ORDER BY b.budget_id DESC
            ''', (session['user_id'],)).fetchall()
            return render_template('index.html', accounts=accounts, current_adviser=current_adviser,
                                   pending_request=pending_request, categories=categories, budgets=budgets)
        if session.get('role') == 'adviser':
            if session.get('is_manager'):
                return redirect(url_for('manager_dashboard'))
            return redirect(url_for('adviser_dashboard'))
        return render_template('index.html')

    @app.route('/accounts/create', methods=['POST'])
    @login_required
    @no_cache
    def create_account():
        if session.get('role') != 'customer':
            return redirect(url_for('index'))

        name = request.form.get('account_name', '').strip()
        acct_type = request.form.get('account_type', 'Checking').strip()
        balance_value = request.form.get('opening_balance', '0').strip() or '0'

        try:
            balance = float(balance_value)
        except ValueError:
            balance = 0.0

        if not name:
            flash('Please provide a valid account name.', 'warning')
            return redirect(url_for('index'))

        db = get_db()
        db.execute(
            'INSERT INTO accounts (customer_id, name, type, balance, currency) VALUES (?, ?, ?, ?, ?)',
            (session['user_id'], name, acct_type, balance, 'GBP')
        )
        db.commit()
        flash('Account created successfully.', 'success')
        return redirect(url_for('index'))

    @app.route('/accounts/<int:account_id>/delete', methods=['POST'])
    @login_required
    @no_cache
    def delete_account(account_id):
        if session.get('role') != 'customer':
            return redirect(url_for('index'))

        db = get_db()
        account = get_customer_account(db, account_id, session['user_id'])
        if not account:
            flash('Account not found.', 'warning')
            return redirect(url_for('index'))

        db.execute('DELETE FROM transactions WHERE account_id = ?', (account_id,))
        db.execute('DELETE FROM accounts WHERE account_id = ?', (account_id,))
        db.commit()
        flash('Account deleted successfully.', 'success')
        return redirect(url_for('index'))

    @app.route('/accounts/<int:account_id>')
    @login_required
    @no_cache
    def account_detail(account_id):
        db = get_db()
        is_adviser_view = False
        back_url = url_for('index')

        if session.get('role') == 'customer':
            account = get_customer_account(db, account_id, session['user_id'])
            if not account:
                flash('Account not found.', 'warning')
                return redirect(url_for('index'))
        elif session.get('role') == 'adviser' and not session.get('is_manager'):
            account = db.execute('SELECT * FROM accounts WHERE account_id = ?', (account_id,)).fetchone()
            if not account:
                flash('Account not found.', 'warning')
                return redirect(url_for('adviser_dashboard'))
            assignment = db.execute(
                "SELECT 1 FROM user_assignments WHERE adviser_id = ? AND customer_id = ? AND status = 'accepted'",
                (session['user_id'], account['customer_id'])
            ).fetchone()
            if not assignment:
                flash('You do not have access to this account.', 'danger')
                return redirect(url_for('adviser_dashboard'))
            if account['is_private']:
                flash('This account is set to private.', 'warning')
                return redirect(url_for('view_client_accounts', customer_id=account['customer_id']))
            is_adviser_view = True
            back_url = url_for('view_client_accounts', customer_id=account['customer_id'])
        else:
            return redirect(url_for('index'))

        owner_id = session['user_id'] if session.get('role') == 'customer' else None
        if owner_id:
            categories = db.execute(
                'SELECT * FROM categories WHERE customer_id IS NULL OR customer_id = ? ORDER BY name',
                (owner_id,)
            ).fetchall()
        else:
            categories = db.execute(
                'SELECT * FROM categories WHERE customer_id IS NULL ORDER BY name'
            ).fetchall()

        transactions = db.execute('''
            SELECT t.*, c.name AS category_name, c.colour AS category_colour
            FROM transactions t
            LEFT JOIN categories c ON t.category_id = c.category_id
            WHERE t.account_id = ?
            ORDER BY t.transaction_date DESC
        ''', (account_id,)).fetchall()

        transactions_asc = db.execute('''
            SELECT t.amount, t.transaction_date, c.name AS category_name, c.colour AS category_colour
            FROM transactions t
            LEFT JOIN categories c ON t.category_id = c.category_id
            WHERE t.account_id = ?
            ORDER BY t.transaction_date ASC, t.transaction_id ASC
        ''', (account_id,)).fetchall()

        # Group by date for line chart
        date_groups = {}
        for t in transactions_asc:
            date = str(t['transaction_date'])[:10]
            amount = t['amount'] or 0
            if date not in date_groups:
                date_groups[date] = {'expenses': 0, 'revenue': 0, 'net': 0}
            if amount < 0:
                date_groups[date]['expenses'] += abs(amount)
            else:
                date_groups[date]['revenue'] += amount
            date_groups[date]['net'] += amount

        total_amounts = sum((t['amount'] or 0) for t in transactions_asc)
        opening_balance = (account['balance'] or 0) - total_amounts

        sorted_dates = sorted(date_groups.keys())
        line_balance, line_expenses, line_revenue = [], [], []
        running = opening_balance
        for date in sorted_dates:
            running += date_groups[date]['net']
            line_balance.append(round(running, 2))
            line_expenses.append(round(date_groups[date]['expenses'], 2))
            line_revenue.append(round(date_groups[date]['revenue'], 2))

        # Group expenses by category for expenditure pie chart
        cat_totals, cat_colours = {}, {}
        for t in transactions_asc:
            amount = t['amount'] or 0
            if amount < 0:
                cat = t['category_name'] or 'Uncategorised'
                cat_totals[cat] = cat_totals.get(cat, 0) + abs(amount)
                cat_colours[cat] = t['category_colour'] or '#AED6F1'

        # Group income by category for income pie chart
        income_totals, income_colours = {}, {}
        for t in transactions_asc:
            amount = t['amount'] or 0
            if amount > 0:
                cat = t['category_name'] or 'Uncategorised'
                income_totals[cat] = income_totals.get(cat, 0) + amount
                income_colours[cat] = t['category_colour'] or '#98D8C8'

        pie_labels = list(cat_totals.keys())
        income_labels = list(income_totals.keys())
        chart_data = json.dumps({
            'line': {
                'labels': sorted_dates,
                'balance': line_balance,
                'expenses': line_expenses,
                'revenue': line_revenue,
            },
            'pie': {
                'labels': pie_labels,
                'amounts': [round(cat_totals[k], 2) for k in pie_labels],
                'colours': [cat_colours[k] for k in pie_labels],
            },
            'pie_income': {
                'labels': income_labels,
                'amounts': [round(income_totals[k], 2) for k in income_labels],
                'colours': [income_colours[k] for k in income_labels],
            }
        })

        goals = []
        if not is_adviser_view:
            goals = db.execute(
                'SELECT * FROM goals WHERE account_id = ? ORDER BY deadline ASC',
                (account_id,)
            ).fetchall()

        return render_template(
            'account_detail.html',
            account=account,
            transactions=transactions,
            categories=categories,
            chart_data=chart_data,
            today=datetime.utcnow().strftime('%Y-%m-%d'),
            is_adviser_view=is_adviser_view,
            back_url=back_url,
            goals=goals
        )

    @app.route('/accounts/<int:account_id>/transactions/add', methods=['POST'])
    @login_required
    @no_cache
    def add_transaction(account_id):
        if session.get('role') != 'customer':
            return redirect(url_for('index'))

        name = request.form.get('transaction_name', '').strip()
        category = request.form.get('category', '').strip()
        description = request.form.get('description', '').strip()
        merchant = request.form.get('merchant', '').strip()
        payment_method = request.form.get('payment_method', '').strip()
        amount_value = request.form.get('amount', '0').strip() or '0'
        transaction_date = request.form.get('transaction_date', '').strip() or datetime.utcnow().strftime('%Y-%m-%d')

        try:
            amount = float(amount_value)
        except ValueError:
            amount = None

        if not name or amount is None:
            flash('Please complete the transaction name and amount.', 'warning')
            return redirect(url_for('account_detail', account_id=account_id))

        db = get_db()
        account = get_customer_account(db, account_id, session['user_id'])
        if not account:
            flash('Account not found.', 'warning')
            return redirect(url_for('index'))

        category_row = db.execute(
            'SELECT category_id FROM categories WHERE name = ? AND (customer_id IS NULL OR customer_id = ?)',
            (category, session['user_id'])
        ).fetchone() if category else None
        category_id = category_row['category_id'] if category_row else None

        db.execute(
            'INSERT INTO transactions (account_id, category_id, name, description, merchant, payment_method, amount, transaction_date) VALUES (?, ?, ?, ?, ?, ?, ?, ?)',
            (account_id, category_id, name, description, merchant or None, payment_method or None, amount, transaction_date)
        )
        db.execute(
            'UPDATE accounts SET balance = balance + ? WHERE account_id = ?',
            (amount, account_id)
        )
        db.commit()
        flash('Transaction added successfully.', 'success')
        return redirect(url_for('account_detail', account_id=account_id))

    @app.route('/accounts/<int:account_id>/transactions/<int:transaction_id>/edit', methods=['POST'])
    @login_required
    @no_cache
    def edit_transaction(account_id, transaction_id):
        if session.get('role') != 'customer':
            return redirect(url_for('index'))

        db = get_db()
        account = get_customer_account(db, account_id, session['user_id'])
        if not account:
            flash('Account not found.', 'warning')
            return redirect(url_for('index'))

        transaction = db.execute(
            'SELECT * FROM transactions WHERE transaction_id = ? AND account_id = ?',
            (transaction_id, account_id)
        ).fetchone()
        if not transaction:
            flash('Transaction not found.', 'warning')
            return redirect(url_for('account_detail', account_id=account_id))

        name = request.form.get('transaction_name', '').strip()
        category = request.form.get('category', '').strip()
        description = request.form.get('description', '').strip()
        merchant = request.form.get('merchant', '').strip()
        payment_method = request.form.get('payment_method', '').strip()
        amount_value = request.form.get('amount', '0').strip() or '0'
        transaction_date = request.form.get('transaction_date', '').strip() or datetime.utcnow().strftime('%Y-%m-%d')

        try:
            new_amount = float(amount_value)
        except ValueError:
            flash('Invalid amount.', 'warning')
            return redirect(url_for('account_detail', account_id=account_id))

        if not name:
            flash('Transaction name is required.', 'warning')
            return redirect(url_for('account_detail', account_id=account_id))

        category_row = db.execute(
            'SELECT category_id FROM categories WHERE name = ? AND (customer_id IS NULL OR customer_id = ?)',
            (category, session['user_id'])
        ).fetchone() if category else None
        category_id = category_row['category_id'] if category_row else None

        old_amount = transaction['amount'] or 0
        balance_delta = new_amount - old_amount

        db.execute('''
            UPDATE transactions
            SET name=?, category_id=?, description=?, merchant=?, payment_method=?, amount=?, transaction_date=?
            WHERE transaction_id=?
        ''', (name, category_id, description, merchant or None, payment_method or None, new_amount, transaction_date, transaction_id))
        db.execute('UPDATE accounts SET balance = balance + ? WHERE account_id = ?', (balance_delta, account_id))
        db.commit()
        flash('Transaction updated successfully.', 'success')
        return redirect(url_for('account_detail', account_id=account_id))

    @app.route('/accounts/<int:account_id>/toggle-privacy', methods=['POST'])
    @login_required
    @no_cache
    def toggle_account_privacy(account_id):
        if session.get('role') != 'customer':
            return redirect(url_for('index'))
        db = get_db()
        account = get_customer_account(db, account_id, session['user_id'])
        if not account:
            flash('Account not found.', 'warning')
            return redirect(url_for('index'))
        db.execute(
            'UPDATE accounts SET is_private = NOT is_private WHERE account_id = ?',
            (account_id,)
        )
        db.commit()
        return redirect(url_for('index'))

    @app.route('/advisers')
    @login_required
    @no_cache
    def browse_advisers():
        if session.get('role') != 'customer':
            return redirect(url_for('index'))
        db = get_db()
        existing_assignment = db.execute(
            'SELECT 1 FROM user_assignments WHERE customer_id = ?', (session['user_id'],)
        ).fetchone()
        if existing_assignment:
            return redirect(url_for('index'))
        pending = db.execute(
            'SELECT request_id FROM customer_requests WHERE customer_id = ?', (session['user_id'],)
        ).fetchone()
        return render_template('advisers.html', pending=pending)

    @app.route('/adviser/request', methods=['POST'])
    @login_required
    @no_cache
    def request_adviser():
        if session.get('role') != 'customer':
            return redirect(url_for('index'))
        db = get_db()
        if db.execute('SELECT 1 FROM user_assignments WHERE customer_id = ?', (session['user_id'],)).fetchone():
            flash('You already have an adviser assigned.', 'warning')
            return redirect(url_for('index'))
        if db.execute('SELECT 1 FROM customer_requests WHERE customer_id = ?', (session['user_id'],)).fetchone():
            flash('You already have a pending request.', 'warning')
            return redirect(url_for('browse_advisers'))
        db.execute('INSERT INTO customer_requests (customer_id) VALUES (?)', (session['user_id'],))
        db.commit()
        flash('Your request has been submitted. A manager will assign you an adviser shortly.', 'success')
        return redirect(url_for('browse_advisers'))

    @app.route('/adviser/cancel-request', methods=['POST'])
    @login_required
    @no_cache
    def cancel_adviser_request():
        if session.get('role') != 'customer':
            return redirect(url_for('index'))
        db = get_db()
        db.execute('DELETE FROM customer_requests WHERE customer_id = ?', (session['user_id'],))
        db.commit()
        flash('Request cancelled.', 'info')
        return redirect(url_for('browse_advisers'))

    @app.route('/adviser/remove', methods=['POST'])
    @login_required
    @no_cache
    def remove_adviser():
        if session.get('role') != 'customer':
            return redirect(url_for('index'))
        db = get_db()
        db.execute('DELETE FROM user_assignments WHERE customer_id = ?', (session['user_id'],))
        db.commit()
        flash('Adviser removed from your account.', 'info')
        return redirect(url_for('index'))

    @app.route('/adviser/dashboard')
    @login_required
    @no_cache
    def adviser_dashboard():
        if session.get('role') != 'adviser' or session.get('is_manager'):
            return redirect(url_for('index'))
        db = get_db()
        clients = db.execute('''
            SELECT c.customer_id, c.name, c.email, c.description
            FROM user_assignments ua
            JOIN customers c ON ua.customer_id = c.customer_id
            WHERE ua.adviser_id = ? AND ua.status = 'accepted'
            ORDER BY c.name
        ''', (session['user_id'],)).fetchall()
        pending_requests = db.execute('''
            SELECT c.customer_id, c.name, c.description
            FROM user_assignments ua
            JOIN customers c ON ua.customer_id = c.customer_id
            WHERE ua.adviser_id = ? AND ua.status = 'pending'
            ORDER BY c.name
        ''', (session['user_id'],)).fetchall()
        tasks = db.execute('''
            SELECT at.*, c.name AS customer_name
            FROM adviser_tasks at
            LEFT JOIN customers c ON at.customer_id = c.customer_id
            WHERE at.adviser_id = ? AND at.status != 'completed'
            ORDER BY at.priority ASC, at.created_at ASC
        ''', (session['user_id'],)).fetchall()
        consultations = db.execute('''
            SELECT con.*, c.name AS customer_name
            FROM consultations con
            JOIN customers c ON con.customer_id = c.customer_id
            WHERE con.adviser_id = ?
            ORDER BY con.scheduled_at DESC, con.created_at DESC
        ''', (session['user_id'],)).fetchall()
        support_tickets = db.execute('''
            SELECT st.*, c.name AS customer_name
            FROM support_tickets st
            JOIN customers c ON st.customer_id = c.customer_id
            JOIN user_assignments ua ON ua.customer_id = st.customer_id
            WHERE ua.adviser_id = ? AND ua.status = 'accepted'
            ORDER BY st.created_at DESC
        ''', (session['user_id'],)).fetchall()
        recent_trend, historical_trend, news = _get_market_data()
        return render_template('adviser_dashboard.html',
                               clients=clients,
                               pending_requests=pending_requests,
                               tasks=tasks,
                               consultations=consultations,
                               support_tickets=support_tickets,
                               recent_trend=recent_trend,
                               historical_trend=historical_trend,
                               news=news)

    @app.route('/adviser/request/accept/<int:customer_id>', methods=['POST'])
    @login_required
    @no_cache
    def accept_request(customer_id):
        if session.get('role') != 'adviser' or session.get('is_manager'):
            return redirect(url_for('index'))
        db = get_db()
        db.execute(
            "UPDATE user_assignments SET status = 'accepted' WHERE adviser_id = ? AND customer_id = ? AND status = 'pending'",
            (session['user_id'], customer_id)
        )
        db.commit()
        flash('Client request accepted.', 'success')
        return redirect(url_for('adviser_dashboard'))

    @app.route('/adviser/request/decline/<int:customer_id>', methods=['POST'])
    @login_required
    @no_cache
    def decline_request(customer_id):
        if session.get('role') != 'adviser' or session.get('is_manager'):
            return redirect(url_for('index'))
        db = get_db()
        db.execute(
            "DELETE FROM user_assignments WHERE adviser_id = ? AND customer_id = ? AND status = 'pending'",
            (session['user_id'], customer_id)
        )
        db.commit()
        flash('Client request declined.', 'info')
        return redirect(url_for('adviser_dashboard'))

    @app.route('/manager/dashboard')
    @login_required
    @no_cache
    def manager_dashboard():
        if session.get('role') != 'adviser' or not session.get('is_manager'):
            return redirect(url_for('index'))
        db = get_db()
        pending_requests = db.execute('''
            SELECT c.customer_id, c.name, c.description
            FROM customer_requests cr
            JOIN customers c ON cr.customer_id = c.customer_id
            ORDER BY c.name
        ''').fetchall()
        advisers = db.execute(
            'SELECT adviser_id, name, description FROM advisers WHERE is_manager = 0 ORDER BY name'
        ).fetchall()
        assignments = db.execute('''
            SELECT c.name AS customer_name, a.name AS adviser_name, c.customer_id, a.adviser_id
            FROM user_assignments ua
            JOIN customers c ON ua.customer_id = c.customer_id
            JOIN advisers a ON ua.adviser_id = a.adviser_id
            ORDER BY c.name
        ''').fetchall()
        support_tickets = db.execute('''
            SELECT st.*, c.name AS customer_name,
                   COALESCE(a.name, 'Unassigned') AS adviser_name
            FROM support_tickets st
            JOIN customers c ON st.customer_id = c.customer_id
            LEFT JOIN user_assignments ua ON ua.customer_id = st.customer_id AND ua.status = 'accepted'
            LEFT JOIN advisers a ON ua.adviser_id = a.adviser_id
            ORDER BY st.status ASC, st.created_at DESC
        ''').fetchall()

        # Performance metrics per adviser
        perf_rows = db.execute('''
            SELECT a.adviser_id, a.name,
                   COUNT(DISTINCT ua.customer_id) AS client_count,
                   COUNT(CASE WHEN at.status != 'completed' THEN 1 END) AS open_tasks,
                   COUNT(CASE WHEN at.status = 'completed' THEN 1 END) AS completed_tasks,
                   AVG(CASE WHEN at.status = 'completed' AND at.completed_at IS NOT NULL
                       THEN (julianday(at.completed_at) - julianday(at.created_at)) * 24
                       ELSE NULL END) AS avg_completion_hours
            FROM advisers a
            LEFT JOIN user_assignments ua ON ua.adviser_id = a.adviser_id AND ua.status = 'accepted'
            LEFT JOIN adviser_tasks at ON at.adviser_id = a.adviser_id
            WHERE a.is_manager = 0
            GROUP BY a.adviser_id, a.name
            ORDER BY a.name
        ''').fetchall()

        # Clients per adviser for roster view
        adviser_clients = {}
        for adv in advisers:
            adviser_clients[adv['adviser_id']] = db.execute('''
                SELECT c.customer_id, c.name, c.email
                FROM user_assignments ua
                JOIN customers c ON ua.customer_id = c.customer_id
                WHERE ua.adviser_id = ? AND ua.status = 'accepted'
                ORDER BY c.name
            ''', (adv['adviser_id'],)).fetchall()

        perf_rows_dicts = [dict(r) for r in perf_rows]

        return render_template('manager_dashboard.html',
                               pending_requests=pending_requests,
                               advisers=advisers,
                               assignments=assignments,
                               support_tickets=support_tickets,
                               perf_rows=perf_rows,
                               perf_rows_json=perf_rows_dicts,
                               adviser_clients=adviser_clients)

    @app.route('/manager/assign/<int:customer_id>', methods=['POST'])
    @login_required
    @no_cache
    def assign_adviser(customer_id):
        if session.get('role') != 'adviser' or not session.get('is_manager'):
            return redirect(url_for('index'))
        adviser_id = request.form.get('adviser_id', type=int)
        if not adviser_id:
            flash('Please select an adviser.', 'warning')
            return redirect(url_for('manager_dashboard'))
        db = get_db()
        adviser = db.execute('SELECT * FROM advisers WHERE adviser_id = ? AND is_manager = 0', (adviser_id,)).fetchone()
        if not adviser:
            flash('Adviser not found.', 'warning')
            return redirect(url_for('manager_dashboard'))
        if not db.execute('SELECT 1 FROM customer_requests WHERE customer_id = ?', (customer_id,)).fetchone():
            flash('No pending request from this customer.', 'warning')
            return redirect(url_for('manager_dashboard'))
        if db.execute('SELECT 1 FROM user_assignments WHERE customer_id = ?', (customer_id,)).fetchone():
            flash('This customer already has an adviser assigned.', 'warning')
            return redirect(url_for('manager_dashboard'))
        db.execute(
            "INSERT INTO user_assignments (adviser_id, customer_id, status) VALUES (?, ?, 'pending')",
            (adviser_id, customer_id)
        )
        db.execute('DELETE FROM customer_requests WHERE customer_id = ?', (customer_id,))
        db.commit()
        flash(f'Adviser assigned to {db.execute("SELECT name FROM customers WHERE customer_id = ?", (customer_id,)).fetchone()["name"]}.', 'success')
        return redirect(url_for('manager_dashboard'))

    @app.route('/manager/unassign/<int:customer_id>', methods=['POST'])
    @login_required
    @no_cache
    def unassign_adviser(customer_id):
        if session.get('role') != 'adviser' or not session.get('is_manager'):
            return redirect(url_for('index'))
        db = get_db()
        db.execute('DELETE FROM user_assignments WHERE customer_id = ?', (customer_id,))
        db.commit()
        flash('Adviser unassigned.', 'info')
        return redirect(url_for('manager_dashboard'))

    @app.route('/adviser/clients/<int:customer_id>')
    @login_required
    @no_cache
    def view_client_accounts(customer_id):
        if session.get('role') != 'adviser' or session.get('is_manager'):
            return redirect(url_for('index'))
        db = get_db()
        assignment = db.execute(
            "SELECT * FROM user_assignments WHERE adviser_id = ? AND customer_id = ? AND status = 'accepted'",
            (session['user_id'], customer_id)
        ).fetchone()
        if not assignment:
            flash('You do not have access to this client.', 'danger')
            return redirect(url_for('adviser_dashboard'))
        customer = db.execute('SELECT * FROM customers WHERE customer_id = ?', (customer_id,)).fetchone()
        accounts = db.execute(
            'SELECT * FROM accounts WHERE customer_id = ? AND is_private = 0 ORDER BY account_id DESC',
            (customer_id,)
        ).fetchall()
        return render_template('client_accounts.html', customer=customer, accounts=accounts)

    @app.route('/accounts/<int:account_id>/transactions/<int:transaction_id>/delete', methods=['POST'])
    @login_required
    @no_cache
    def delete_transaction(account_id, transaction_id):
        if session.get('role') != 'customer':
            return redirect(url_for('index'))

        db = get_db()
        account = get_customer_account(db, account_id, session['user_id'])
        if not account:
            flash('Account not found.', 'warning')
            return redirect(url_for('index'))

        transaction = db.execute(
            'SELECT * FROM transactions WHERE transaction_id = ? AND account_id = ?',
            (transaction_id, account_id)
        ).fetchone()
        if not transaction:
            flash('Transaction not found.', 'warning')
            return redirect(url_for('account_detail', account_id=account_id))

        db.execute('DELETE FROM transactions WHERE transaction_id = ?', (transaction_id,))
        db.execute('UPDATE accounts SET balance = balance - ? WHERE account_id = ?', (transaction['amount'], account_id))
        db.commit()
        flash('Transaction removed successfully.', 'success')
        return redirect(url_for('account_detail', account_id=account_id))

    @app.route('/accounts/<int:account_id>/goals/add', methods=['POST'])
    @login_required
    @no_cache
    def add_goal(account_id):
        if session.get('role') != 'customer':
            return redirect(url_for('index'))

        db = get_db()
        account = get_customer_account(db, account_id, session['user_id'])
        if not account:
            flash('Account not found.', 'warning')
            return redirect(url_for('index'))

        name          = request.form.get('goal_name', '').strip()
        target_str    = request.form.get('target_amount', '').strip()
        deadline      = request.form.get('deadline', '').strip() or None

        if not name or not target_str:
            flash('Goal name and target amount are required.', 'warning')
            return redirect(url_for('account_detail', account_id=account_id))

        try:
            target = float(target_str)
            if target <= 0:
                raise ValueError
        except ValueError:
            flash('Target amount must be a positive number.', 'warning')
            return redirect(url_for('account_detail', account_id=account_id))

        db.execute(
            'INSERT INTO goals (account_id, name, target_amount, deadline) VALUES (?, ?, ?, ?)',
            (account_id, name, target, deadline)
        )
        db.commit()
        flash('Goal added.', 'success')
        return redirect(url_for('account_detail', account_id=account_id))

    @app.route('/accounts/<int:account_id>/goals/<int:goal_id>/delete', methods=['POST'])
    @login_required
    @no_cache
    def delete_goal(account_id, goal_id):
        if session.get('role') != 'customer':
            return redirect(url_for('index'))

        db = get_db()
        account = get_customer_account(db, account_id, session['user_id'])
        if not account:
            flash('Account not found.', 'warning')
            return redirect(url_for('index'))

        db.execute('DELETE FROM goals WHERE goal_id = ? AND account_id = ?', (goal_id, account_id))
        db.commit()
        flash('Goal removed.', 'success')
        return redirect(url_for('account_detail', account_id=account_id))

    @app.route('/budgets/add', methods=['POST'])
    @login_required
    @no_cache
    def add_budget():
        if session.get('role') != 'customer':
            return redirect(url_for('index'))

        category_id = request.form.get('category_id', '').strip()
        max_str     = request.form.get('maximum_amount', '').strip()
        start_date  = request.form.get('start_date', '').strip() or None
        end_date    = request.form.get('end_date', '').strip() or None

        if not category_id or not max_str:
            flash('Category and maximum amount are required.', 'warning')
            return redirect(url_for('index'))

        try:
            max_amount = float(max_str)
            if max_amount <= 0:
                raise ValueError
        except ValueError:
            flash('Maximum amount must be a positive number.', 'warning')
            return redirect(url_for('index'))

        db = get_db()
        # Prevent duplicate budget for the same category in overlapping period
        existing = db.execute(
            'SELECT 1 FROM budgets WHERE customer_id = ? AND category_id = ?',
            (session['user_id'], int(category_id))
        ).fetchone()
        if existing:
            flash('You already have a budget for that category. Delete it first.', 'warning')
            return redirect(url_for('index'))

        db.execute(
            'INSERT INTO budgets (customer_id, category_id, maximum_amount, start_date, end_date) VALUES (?, ?, ?, ?, ?)',
            (session['user_id'], int(category_id), max_amount, start_date, end_date)
        )
        db.commit()
        flash('Budget created.', 'success')
        return redirect(url_for('index'))

    @app.route('/budgets/<int:budget_id>/delete', methods=['POST'])
    @login_required
    @no_cache
    def delete_budget(budget_id):
        if session.get('role') != 'customer':
            return redirect(url_for('index'))

        db = get_db()
        db.execute('DELETE FROM budgets WHERE budget_id = ? AND customer_id = ?', (budget_id, session['user_id']))
        db.commit()
        flash('Budget removed.', 'success')
        return redirect(url_for('index'))

    @app.route('/support/submit', methods=['POST'])
    @login_required
    @no_cache
    def submit_support_ticket():
        if session.get('role') != 'customer':
            return redirect(url_for('index'))

        ticket_type = request.form.get('ticket_type', 'help')
        subject = request.form.get('subject', '').strip()
        message = request.form.get('message', '').strip()

        if not subject or not message:
            flash('Please fill in all fields.', 'warning')
            return redirect(url_for('index'))

        db = get_db()
        db.execute(
            'INSERT INTO support_tickets (customer_id, type, subject, message) VALUES (?, ?, ?, ?)',
            (session['user_id'], ticket_type, subject, message)
        )
        ticket_id = db.execute('SELECT last_insert_rowid()').fetchone()[0]

        assignment = db.execute(
            "SELECT adviser_id FROM user_assignments WHERE customer_id = ? AND status = 'accepted'",
            (session['user_id'],)
        ).fetchone()

        if assignment:
            priority = 1 if ticket_type == 'complaint' else 2
            customer = db.execute('SELECT name FROM customers WHERE customer_id = ?', (session['user_id'],)).fetchone()
            task_title = f"{'Complaint' if ticket_type == 'complaint' else 'Help request'}: {subject}"
            db.execute('''
                INSERT INTO adviser_tasks (adviser_id, customer_id, ticket_id, title, description, priority)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (assignment['adviser_id'], session['user_id'], ticket_id, task_title,
                  f"From {customer['name']}: {message}", priority))

        db.commit()
        flash('Your request has been submitted.', 'success')
        return redirect(url_for('index'))

    @app.route('/support/tickets/<int:ticket_id>/resolve', methods=['POST'])
    @login_required
    @no_cache
    def resolve_ticket(ticket_id):
        if session.get('role') not in ('adviser',) and not session.get('is_manager'):
            return redirect(url_for('index'))
        db = get_db()
        db.execute(
            "UPDATE support_tickets SET status='resolved', resolved_at=CURRENT_TIMESTAMP WHERE ticket_id=?",
            (ticket_id,)
        )
        db.commit()
        flash('Ticket marked as resolved.', 'success')
        return redirect(request.referrer or url_for('adviser_dashboard'))

    @app.route('/consultations/create', methods=['POST'])
    @login_required
    @no_cache
    def create_consultation():
        if session.get('role') != 'customer':
            return redirect(url_for('index'))

        db = get_db()
        assignment = db.execute(
            "SELECT adviser_id FROM user_assignments WHERE customer_id = ? AND status = 'accepted'",
            (session['user_id'],)
        ).fetchone()
        if not assignment:
            flash('You need an assigned adviser before scheduling a consultation.', 'warning')
            return redirect(url_for('index'))

        title = request.form.get('title', '').strip()
        notes = request.form.get('notes', '').strip()
        scheduled_at = request.form.get('scheduled_at', '').strip() or None

        if not title:
            flash('Please provide a consultation title.', 'warning')
            return redirect(url_for('index'))

        db.execute('''
            INSERT INTO consultations (customer_id, adviser_id, title, notes, scheduled_at)
            VALUES (?, ?, ?, ?, ?)
        ''', (session['user_id'], assignment['adviser_id'], title, notes or None, scheduled_at))
        db.commit()
        flash('Consultation request submitted.', 'success')
        return redirect(url_for('index'))

    @app.route('/consultations/<int:consultation_id>/update', methods=['POST'])
    @login_required
    @no_cache
    def update_consultation(consultation_id):
        if session.get('role') != 'adviser' or session.get('is_manager'):
            return redirect(url_for('index'))
        new_status = request.form.get('status', 'scheduled')
        db = get_db()
        db.execute(
            'UPDATE consultations SET status=? WHERE consultation_id=? AND adviser_id=?',
            (new_status, consultation_id, session['user_id'])
        )
        db.commit()
        flash('Consultation updated.', 'success')
        return redirect(url_for('adviser_dashboard'))

    @app.route('/adviser/clients/<int:customer_id>/export')
    @login_required
    @no_cache
    def export_client_transactions(customer_id):
        if session.get('role') != 'adviser' or session.get('is_manager'):
            return redirect(url_for('index'))
        db = get_db()
        assignment = db.execute(
            "SELECT 1 FROM user_assignments WHERE adviser_id=? AND customer_id=? AND status='accepted'",
            (session['user_id'], customer_id)
        ).fetchone()
        if not assignment:
            flash('You do not have access to this client.', 'danger')
            return redirect(url_for('adviser_dashboard'))

        customer = db.execute('SELECT name FROM customers WHERE customer_id=?', (customer_id,)).fetchone()
        rows = db.execute('''
            SELECT a.name AS account_name, t.transaction_date, t.name, cat.name AS category,
                   t.merchant, t.payment_method, t.description, t.amount
            FROM transactions t
            JOIN accounts a ON t.account_id = a.account_id
            LEFT JOIN categories cat ON t.category_id = cat.category_id
            WHERE a.customer_id = ? AND a.is_private = 0
            ORDER BY t.transaction_date DESC
        ''', (customer_id,)).fetchall()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Transactions'

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='2563EB')
        headers = ['Date', 'Account', 'Transaction Name', 'Category', 'Merchant', 'Payment Method', 'Description', 'Amount (£)']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for row_idx, row in enumerate(rows, 2):
            ws.cell(row=row_idx, column=1, value=str(row['transaction_date'])[:10])
            ws.cell(row=row_idx, column=2, value=row['account_name'])
            ws.cell(row=row_idx, column=3, value=row['name'])
            ws.cell(row=row_idx, column=4, value=row['category'] or '')
            ws.cell(row=row_idx, column=5, value=row['merchant'] or '')
            ws.cell(row=row_idx, column=6, value=row['payment_method'] or '')
            ws.cell(row=row_idx, column=7, value=row['description'] or '')
            amount_cell = ws.cell(row=row_idx, column=8, value=row['amount'] or 0)
            if (row['amount'] or 0) < 0:
                amount_cell.font = Font(color='CC0000')
            else:
                amount_cell.font = Font(color='006600')

        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"{customer['name'].replace(' ', '_')}_transactions.xlsx"
        return send_file(buf, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/adviser/tasks/<int:task_id>/complete', methods=['POST'])
    @login_required
    @no_cache
    def complete_task(task_id):
        if session.get('role') != 'adviser' or session.get('is_manager'):
            return redirect(url_for('index'))
        db = get_db()
        db.execute('''
            UPDATE adviser_tasks SET status='completed', completed_at=CURRENT_TIMESTAMP
            WHERE task_id=? AND adviser_id=?
        ''', (task_id, session['user_id']))
        db.commit()
        flash('Task marked as complete.', 'success')
        return redirect(url_for('adviser_dashboard'))

    @app.route('/manager/reassign/<int:customer_id>', methods=['POST'])
    @login_required
    @no_cache
    def reassign_client(customer_id):
        if session.get('role') != 'adviser' or not session.get('is_manager'):
            return redirect(url_for('index'))
        new_adviser_id = request.form.get('new_adviser_id', type=int)
        if not new_adviser_id:
            flash('Please select a new adviser.', 'warning')
            return redirect(url_for('manager_dashboard'))
        db = get_db()
        adviser = db.execute('SELECT * FROM advisers WHERE adviser_id=? AND is_manager=0', (new_adviser_id,)).fetchone()
        if not adviser:
            flash('Adviser not found.', 'warning')
            return redirect(url_for('manager_dashboard'))
        db.execute(
            "UPDATE user_assignments SET adviser_id=?, status='accepted' WHERE customer_id=?",
            (new_adviser_id, customer_id)
        )
        db.commit()
        flash('Client reassigned successfully.', 'success')
        return redirect(url_for('manager_dashboard'))

    @app.route('/register', methods=['GET', 'POST'])
    def register():
        if session.get('username'):
            return redirect(url_for('index'))

        if request.method == 'POST':
            name     = request.form.get('name', '').strip()
            email    = request.form.get('email', '').strip().lower()
            password = request.form.get('password', '')
            confirm  = request.form.get('confirm_password', '')

            if not name or not email or not password:
                flash('All fields are required.', 'danger')
                return render_template('register.html', form_data=request.form)

            if password != confirm:
                flash('Passwords do not match.', 'danger')
                return render_template('register.html', form_data=request.form)

            if len(password) < 8:
                flash('Password must be at least 8 characters.', 'danger')
                return render_template('register.html', form_data=request.form)

            db = get_db()
            existing = db.execute('SELECT 1 FROM customers WHERE email = ?', (email,)).fetchone()
            if existing:
                flash('An account with that email already exists.', 'danger')
                return render_template('register.html', form_data=request.form)

            db.execute(
                'INSERT INTO customers (name, email, password, currency) VALUES (?, ?, ?, ?)',
                (name, email, generate_password_hash(password), 'GBP')
            )
            db.commit()

            flash('Account created successfully. Please sign in.', 'success')
            return redirect(url_for('login'))

        return render_template('register.html')

    @app.route('/login', methods=['GET', 'POST'])
    def login():
        if request.method == 'POST':
            db = get_db()
            user_type = request.form.get('user_type', 'customer')
            email = request.form.get('email', '').strip()
            password = request.form.get('password', '')
            user = None
            role = None

            if user_type == 'adviser':
                user = db.execute('SELECT * FROM advisers WHERE email = ?', (email,)).fetchone()
                role = 'adviser'
            else:
                user = db.execute('SELECT * FROM customers WHERE email = ?', (email,)).fetchone()
                role = 'customer'

            if user and check_password_hash(user['password'], password):
                pending_data = {
                    'user_id': user['adviser_id'] if role == 'adviser' else user['customer_id'],
                    'username': user['name'],
                    'role': role,
                    'is_manager': bool(user['is_manager']) if role == 'adviser' else False
                }

                if user['is_2fa_enabled']:
                    session['pending_2fa'] = pending_data
                    return redirect(url_for('two_factor'))

                session.update(pending_data)
                if role == 'adviser' and pending_data['is_manager']:
                    flash('Welcome manager. Your adviser account has is_manager access.', 'success')
                return redirect(url_for('index'))

            flash('Invalid email or password. Please try again.', 'danger')
        return render_template('login.html')

    @app.route('/two-factor', methods=['GET', 'POST'])
    @no_cache
    def two_factor():
        pending = session.get('pending_2fa')
        if not pending:
            flash('Please sign in before entering a 2FA code.', 'warning')
            return redirect(url_for('login'))

        if request.method == 'POST':
            code = request.form.get('code', '').strip()
            db = get_db()
            user = get_user_by_role(db, pending['role'], pending['user_id'])

            if user and user['totp_secret'] and pyotp.TOTP(user['totp_secret']).verify(code, valid_window=1):
                session.update(pending)
                session.pop('pending_2fa', None)
                flash('Sign in successful.', 'success')
                return redirect(url_for('index'))

            flash('Invalid authentication code. Please try again.', 'danger')

        return render_template('two_factor.html')

    @app.route('/setup-2fa', methods=['GET', 'POST'])
    @login_required
    @no_cache
    def setup_2fa():
        db = get_db()
        user = get_user_by_role(db, session['role'], session['user_id'])

        if not user:
            session.clear()
            return redirect(url_for('login'))

        if request.method == 'POST':
            code = request.form.get('code', '').strip()
            secret = session.get('totp_secret_setup')

            if not secret:
                flash('Your 2FA setup session expired. Please refresh the page.', 'warning')
                return redirect(url_for('setup_2fa'))

            if pyotp.TOTP(secret).verify(code, valid_window=1):
                if session['role'] == 'adviser':
                    db.execute('UPDATE advisers SET totp_secret = ?, is_2fa_enabled = 1 WHERE adviser_id = ?', (secret, session['user_id']))
                else:
                    db.execute('UPDATE customers SET totp_secret = ?, is_2fa_enabled = 1 WHERE customer_id = ?', (secret, session['user_id']))
                db.commit()
                session.pop('totp_secret_setup', None)
                flash('Two-factor authentication is now enabled for your account.', 'success')
                return redirect(url_for('index'))

            flash('That code is not valid. Please try again.', 'danger')

        if user['is_2fa_enabled']:
            flash('Two-factor authentication is already enabled for your account.', 'info')
            return redirect(url_for('index'))

        secret = pyotp.random_base32()
        session['totp_secret_setup'] = secret
        provisioning_uri = pyotp.totp.TOTP(secret).provisioning_uri(
            name=user['email'],
            issuer_name='DWK Finance'
        )

        qr = qrcode.QRCode(box_size=8, border=2)
        qr.add_data(provisioning_uri)
        qr.make(fit=True)
        img = qr.make_image(fill_color='black', back_color='white')

        buffer = io.BytesIO()
        img.save(buffer, format='PNG')
        qr_code_data = base64.b64encode(buffer.getvalue()).decode()

        return render_template(
            'setup_2fa.html',
            qr_code_data=qr_code_data,
            secret=secret
        )

    @app.route('/logout')
    def logout():
        session.clear()
        flash('You have been logged out.', 'success')
        return redirect(url_for('login'))

    @app.route('/forgot-password', methods=['GET', 'POST'])
    def forgot_password():
        """Handle password reset requests"""
        if request.method == 'POST':
            email = request.form.get('email', '').strip()
            db = get_db()
            
            #Check if email exists in either customers or advisers table
            customer = db.execute('SELECT * FROM customers WHERE email = ?', (email,)).fetchone()
            adviser = db.execute('SELECT * FROM advisers WHERE email = ?', (email,)).fetchone()
            
            if customer or adviser:
                # TODO: In production, send a password reset email here
                # For now, we'll just show a success message
                flash(
                    'If an account exists with that email, you will receive password reset instructions shortly. '
                    'Please check your email and follow the instructions provided.',
                    'success'
                )
            else:
                #show success message for security reasons
                flash(
                    'If an account exists with that email, you will receive password reset instructions shortly. '
                    'Please check your email and follow the instructions provided.',
                    'info'
                )
            
            return redirect(url_for('login'))
        
        return render_template('forgot_password.html')

    @app.route('/about-vectura')
    def about_vectura():
        """Serve the Vectura team information page"""
        return render_template('about_vectura.html')

    @app.route('/terms')
    def terms():
        """Serve the Terms of Service page"""
        return render_template('terms.html')

    @app.route('/privacy')
    def privacy():
        """Serve the Privacy Policy page"""
        return render_template('privacy.html')

    @app.route('/settings', methods=['GET', 'POST'])
    @login_required
    @no_cache
    def settings():
        db = get_db()
        user = get_user_by_role(db, session['role'], session['user_id'])

        if not user:
            session.clear()
            return redirect(url_for('login'))

        if request.method == 'POST':
            action = request.form.get('action')

            if action == 'update_profile':
                name  = request.form.get('name', '').strip()
                email = request.form.get('email', '').strip().lower()

                if not name or not email:
                    flash('Name and email are required.', 'danger')
                    return redirect(url_for('settings'))

                # Check email is not already taken by another account
                if session['role'] == 'adviser':
                    clash = db.execute(
                        'SELECT 1 FROM advisers WHERE email = ? AND adviser_id != ?',
                        (email, session['user_id'])
                    ).fetchone()
                else:
                    clash = db.execute(
                        'SELECT 1 FROM customers WHERE email = ? AND customer_id != ?',
                        (email, session['user_id'])
                    ).fetchone()

                if clash:
                    flash('That email address is already in use.', 'danger')
                    return redirect(url_for('settings'))

                if session['role'] == 'adviser':
                    db.execute('UPDATE advisers SET name = ?, email = ? WHERE adviser_id = ?',
                               (name, email, session['user_id']))
                else:
                    currency = request.form.get('currency', user['currency'])
                    db.execute('UPDATE customers SET name = ?, email = ?, currency = ? WHERE customer_id = ?',
                               (name, email, currency, session['user_id']))

                db.commit()
                session['username'] = name
                flash('Profile updated successfully.', 'success')
                return redirect(url_for('settings'))

            elif action == 'change_password':
                current  = request.form.get('current_password', '')
                new_pw   = request.form.get('new_password', '')
                confirm  = request.form.get('confirm_password', '')

                if not check_password_hash(user['password'], current):
                    flash('Current password is incorrect.', 'danger')
                    return redirect(url_for('settings'))

                if new_pw != confirm:
                    flash('New passwords do not match.', 'danger')
                    return redirect(url_for('settings'))

                if len(new_pw) < 8:
                    flash('New password must be at least 8 characters.', 'danger')
                    return redirect(url_for('settings'))

                if session['role'] == 'adviser':
                    db.execute('UPDATE advisers SET password = ? WHERE adviser_id = ?',
                               (generate_password_hash(new_pw), session['user_id']))
                else:
                    db.execute('UPDATE customers SET password = ? WHERE customer_id = ?',
                               (generate_password_hash(new_pw), session['user_id']))

                db.commit()
                flash('Password changed successfully.', 'success')
                return redirect(url_for('settings'))

            elif action == 'disable_2fa':
                code = request.form.get('code', '').strip()
                if user['totp_secret'] and pyotp.TOTP(user['totp_secret']).verify(code, valid_window=1):
                    if session['role'] == 'adviser':
                        db.execute('UPDATE advisers SET totp_secret = NULL, is_2fa_enabled = 0 WHERE adviser_id = ?', (session['user_id'],))
                    else:
                        db.execute('UPDATE customers SET totp_secret = NULL, is_2fa_enabled = 0 WHERE customer_id = ?', (session['user_id'],))
                    db.commit()
                    flash('Two-factor authentication has been disabled.', 'success')
                    return redirect(url_for('settings'))
                else:
                    flash('Invalid authentication code. Please try again.', 'danger')

        custom_categories = []
        if session.get('role') == 'customer':
            custom_categories = db.execute(
                'SELECT * FROM categories WHERE customer_id = ? ORDER BY name',
                (session['user_id'],)
            ).fetchall()

        return render_template('settings.html', user=user, custom_categories=custom_categories)

    @app.route('/categories/add', methods=['POST'])
    @login_required
    @no_cache
    def add_category():
        if session.get('role') != 'customer':
            return redirect(url_for('settings'))

        name   = request.form.get('cat_name', '').strip()
        colour = request.form.get('cat_colour', '#AED6F1').strip()

        if not name:
            flash('Category name is required.', 'warning')
            return redirect(url_for('settings'))

        db = get_db()
        existing = db.execute(
            'SELECT 1 FROM categories WHERE name = ? AND (customer_id IS NULL OR customer_id = ?)',
            (name, session['user_id'])
        ).fetchone()
        if existing:
            flash(f'A category named "{name}" already exists.', 'warning')
            return redirect(url_for('settings'))

        db.execute(
            'INSERT INTO categories (name, colour, customer_id) VALUES (?, ?, ?)',
            (name, colour, session['user_id'])
        )
        db.commit()
        flash(f'Category "{name}" added.', 'success')
        return redirect(url_for('settings'))

    @app.route('/categories/<int:category_id>/edit', methods=['POST'])
    @login_required
    @no_cache
    def edit_category(category_id):
        if session.get('role') != 'customer':
            return redirect(url_for('settings'))

        db = get_db()
        cat = db.execute(
            'SELECT * FROM categories WHERE category_id = ? AND customer_id = ?',
            (category_id, session['user_id'])
        ).fetchone()
        if not cat:
            flash('Category not found.', 'warning')
            return redirect(url_for('settings'))

        name   = request.form.get('cat_name', '').strip()
        colour = request.form.get('cat_colour', cat['colour']).strip()

        if not name:
            flash('Category name is required.', 'warning')
            return redirect(url_for('settings'))

        clash = db.execute(
            'SELECT 1 FROM categories WHERE name = ? AND (customer_id IS NULL OR customer_id = ?) AND category_id != ?',
            (name, session['user_id'], category_id)
        ).fetchone()
        if clash:
            flash(f'A category named "{name}" already exists.', 'warning')
            return redirect(url_for('settings'))

        db.execute(
            'UPDATE categories SET name = ?, colour = ? WHERE category_id = ?',
            (name, colour, category_id)
        )
        db.commit()
        flash('Category updated.', 'success')
        return redirect(url_for('settings'))

    @app.route('/categories/<int:category_id>/delete', methods=['POST'])
    @login_required
    @no_cache
    def delete_category(category_id):
        if session.get('role') != 'customer':
            return redirect(url_for('settings'))

        db = get_db()
        cat = db.execute(
            'SELECT * FROM categories WHERE category_id = ? AND customer_id = ?',
            (category_id, session['user_id'])
        ).fetchone()
        if not cat:
            flash('Category not found.', 'warning')
            return redirect(url_for('settings'))

        # Nullify transactions and remove budgets using this category
        db.execute('UPDATE transactions SET category_id = NULL WHERE category_id = ?', (category_id,))
        db.execute('DELETE FROM budgets WHERE category_id = ? AND customer_id = ?', (category_id, session['user_id']))
        db.execute('DELETE FROM categories WHERE category_id = ?', (category_id,))
        db.commit()
        flash(f'Category "{cat["name"]}" deleted.', 'success')
        return redirect(url_for('settings'))

    @app.route('/contact')
    def contact():
        """Serve the Contact Us page"""
        return render_template('contact.html')
